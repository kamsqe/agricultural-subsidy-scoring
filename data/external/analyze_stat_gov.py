"""
Analyze stat.gov.kz datasets and create feature-ready regional tables.
Outputs: data/external/regional_features.csv
"""
import os
import json
import csv
import warnings
warnings.filterwarnings('ignore')

DATA_DIR = os.path.dirname(os.path.abspath(__file__)) + '/stat_gov_kz'
OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# 1. Parse element_458518.xlsx — comprehensive livestock census
# ============================================================
import xlrd

CENSUS_FILE = os.path.join(DATA_DIR, 'element_458518.xlsx')
wb = xlrd.open_workbook(CENSUS_FILE)

# Mapping of sheet names to what we extract
SHEETS = {
    'КРС (регионы)': 'cattle',
    'Овцы (регионы)': 'sheep',
    'Козы (регионы)': 'goats',
    'Лошади (регионы)': 'horses',
    'Верблюды (регионы)': 'camels',
    'Свиньи (регионы)': 'pigs',
    'Птица (регионы)': 'poultry',
    'МОЛОКО (регионы)': 'milk',
    'МЯСО в живом весе (регионы)': 'meat_live',
}

def normalize_region(name):
    """Normalize region name to match our dataset"""
    name = str(name).strip()
    # Remove "область" suffix for matching
    mappings = {
        'Абай': 'abai',
        'Акмолинская': 'akmolinsk',
        'Актюбинская': 'aktobe',
        'Алматинская': 'almaty_obl',
        'Атырауская': 'atyrau',
        'Западно-Казахстанская': 'west_kz',
        'Жамбылская': 'zhambyl',
        'Жетісу': 'zhetysu',
        'Карагандинская': 'karaganda',
        'Костанайская': 'kostanay',
        'Кызылординская': 'kyzylorda',
        'Мангистауская': 'mangystau',
        'Павлодарская': 'pavlodar',
        'Северо-Казахстанская': 'north_kz',
        'Туркестанская': 'turkestan',
        'Ұлытау': 'ulytau',
        'Восточно-Казахстанская': 'east_kz',
        'г.Астана': 'astana',
        'г.Алматы': 'almaty_city',
        'г.Шымкент': 'shymkent',
        'Республика Казахстан': 'RK_TOTAL',
    }
    for key, val in mappings.items():
        if key.lower() in name.lower():
            return val
    return None

def extract_regional_data(sheet_name, metric_name):
    """Extract latest year data by region from a sheet"""
    sheet = wb.sheet_by_name(sheet_name)
    
    # Find year row (usually row 3 or 4)
    year_row = None
    for r in range(min(6, sheet.nrows)):
        for c in range(1, min(40, sheet.ncols)):
            v = sheet.cell_value(r, c)
            if isinstance(v, (int, float)) and 1990 <= v <= 2030:
                year_row = r
                break
        if year_row is not None:
            break
    
    if year_row is None:
        print(f"  WARNING: No year row found in {sheet_name}")
        return {}
    
    # Find last year column with data
    years = {}
    for c in range(1, sheet.ncols):
        v = sheet.cell_value(year_row, c)
        if isinstance(v, (int, float)) and 1990 <= v <= 2030:
            years[c] = int(v)
    
    if not years:
        return {}
    
    # Get latest 3 years
    sorted_cols = sorted(years.items(), key=lambda x: x[1], reverse=True)
    latest_cols = sorted_cols[:3]  # [2024, 2023, 2022]
    
    results = {}
    # Find data rows (start after year row)
    for r in range(year_row + 1, sheet.nrows):
        region_name = str(sheet.cell_value(r, 0)).strip()
        if not region_name:
            continue
        
        region_key = normalize_region(region_name)
        if not region_key or region_key == 'RK_TOTAL':
            continue
        
        for col, year in latest_cols:
            v = sheet.cell_value(r, col)
            if isinstance(v, (int, float)) and v > 0:
                key = f'{metric_name}_{year}'
                if region_key not in results:
                    results[region_key] = {}
                results[region_key][key] = round(v, 2)
                break  # Use latest available year
    
    return results

# Extract all metrics
print("Extracting regional data from stat.gov.kz census...")
all_regions = {}

for sheet_name, metric in SHEETS.items():
    try:
        data = extract_regional_data(sheet_name, metric)
        for region, values in data.items():
            if region not in all_regions:
                all_regions[region] = {'region': region}
            all_regions[region].update(values)
        print(f"  {sheet_name}: {len(data)} regions")
    except Exception as e:
        print(f"  ERROR {sheet_name}: {e}")

# xlrd Book doesn't have close()

# ============================================================
# 2. Parse farm grouping files (440925, 440926)
# ============================================================
import openpyxl

def extract_farm_grouping(element_id, animal_type):
    """Extract farm size distribution by region"""
    fpath = os.path.join(DATA_DIR, f'element_{element_id}.xlsx')
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    sheet = wb2.active
    
    results = {}
    for r_idx, row in enumerate(sheet.iter_rows(min_row=9, values_only=True)):
        region_name = str(row[0]).strip() if row[0] else ''
        region_key = normalize_region(region_name)
        if not region_key or region_key == 'RK_TOTAL':
            continue
        
        try:
            total_farms = int(row[1]) if row[1] else 0
            small_farms = int(row[2]) if row[2] else 0  # <10 heads
            medium_farms = int(row[4]) if row[4] else 0  # 11-50 heads
            large_farms = (int(row[6]) if row[6] else 0) + (int(row[8]) if row[8] else 0)  # 51-500
            
            results[region_key] = {
                f'{animal_type}_total_farms': total_farms,
                f'{animal_type}_small_pct': round(small_farms / max(total_farms, 1) * 100, 1),
                f'{animal_type}_medium_pct': round(medium_farms / max(total_farms, 1) * 100, 1),
            }
        except (ValueError, TypeError):
            pass
    
    wb2.close()
    return results

print("\nExtracting farm grouping data...")
for eid, animal in [(440925, 'cattle'), (440926, 'sheep')]:
    try:
        data = extract_farm_grouping(eid, animal)
        for region, values in data.items():
            if region not in all_regions:
                all_regions[region] = {'region': region}
            all_regions[region].update(values)
        print(f"  element_{eid} ({animal}): {len(data)} regions")
    except Exception as e:
        print(f"  ERROR element_{eid}: {e}")

# ============================================================
# 3. Parse monthly data for latest trends (8126=cattle, 8128=sheep)
# ============================================================
def extract_monthly_trend(element_id, metric):
    """Extract YoY growth from monthly data"""
    fpath = os.path.join(DATA_DIR, f'element_{element_id}.xlsx')
    wb3 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    sheet = wb3.active
    
    # Structure: rows are regions, columns are months across years
    # Try to get Jan values for latest 2 years to compute YoY
    results = {}
    rows = list(sheet.iter_rows(values_only=True))
    
    # Find data start row (after headers)
    data_start = 4  # typical
    
    for r_idx in range(data_start, len(rows)):
        row = rows[r_idx]
        if not row[0]:
            continue
        region_name = str(row[0]).strip()
        region_key = normalize_region(region_name)
        if not region_key or region_key == 'RK_TOTAL':
            continue
        
        # Get last two non-None values (latest months)
        values = [v for v in row[1:] if isinstance(v, (int, float)) and v > 0]
        if len(values) >= 2:
            latest = values[-1]
            previous = values[-2]
            if previous > 0:
                yoy = round((latest - previous) / previous * 100, 2)
                results[region_key] = {f'{metric}_yoy_pct': yoy, f'{metric}_latest': round(latest, 2)}
    
    wb3.close()
    return results

print("\nExtracting monthly trends...")
for eid, metric in [(8126, 'cattle_monthly'), (8128, 'sheep_monthly')]:
    try:
        data = extract_monthly_trend(eid, metric)
        for region, values in data.items():
            if region not in all_regions:
                all_regions[region] = {'region': region}
            all_regions[region].update(values)
        print(f"  element_{eid} ({metric}): {len(data)} regions")
    except Exception as e:
        print(f"  ERROR element_{eid}: {e}")

# ============================================================
# 4. Save results
# ============================================================
print(f"\n=== RESULTS ===")
print(f"Total regions: {len(all_regions)}")

# Collect all column names
all_cols = set()
for r in all_regions.values():
    all_cols.update(r.keys())
all_cols = sorted(all_cols)

print(f"Total features: {len(all_cols)}")
print(f"Columns: {all_cols}")

# Save as CSV
out_path = os.path.join(OUT_DIR, 'regional_features.csv')
with open(out_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=all_cols)
    writer.writeheader()
    for region_data in sorted(all_regions.values(), key=lambda x: x['region']):
        writer.writerow(region_data)

print(f"\nSaved to: {out_path}")

# Also save as JSON for easy inspection
json_path = os.path.join(OUT_DIR, 'regional_features.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(all_regions, f, ensure_ascii=False, indent=2)
print(f"Saved to: {json_path}")

# Print sample
print("\n=== SAMPLE (Акмолинская) ===")
sample = all_regions.get('akmolinsk', {})
for k, v in sorted(sample.items()):
    print(f"  {k}: {v}")
